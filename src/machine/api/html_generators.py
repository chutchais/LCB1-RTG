# def json_to_html_table(report_data):
#     """
#     Convert productivity report JSON to HTML table format
    
#     Args:
#         report_data: Dict from get_productivity_report_daily response
    
#     Returns:
#         HTML string of formatted table
#     """
    
#     if not report_data.get('daily_reports') or len(report_data['daily_reports']) == 0:
#         return '<p>No data available</p>'
    
#     html = '<table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">\n'
    
#     # Header row
#     html += '<thead>\n<tr style="background-color: #2c3e50; color: white;">\n'
#     html += '  <th>Date</th>\n'
#     html += '  <th>Equipment</th>\n'
#     html += '  <th colspan="3">Morning (08:00-20:00)</th>\n'
#     html += '  <th colspan="3">Night (20:00-08:00)</th>\n'
#     html += '  <th colspan="3">Total</th>\n'
#     html += '</tr>\n<tr style="background-color: #2c3e50; color: white;">\n'
#     html += '  <th></th>\n'
#     html += '  <th></th>\n'
#     html += '  <th>Hour</th>\n'
#     html += '  <th>Move</th>\n'
#     html += '  <th>Prod</th>\n'
#     html += '  <th>Hour</th>\n'
#     html += '  <th>Move</th>\n'
#     html += '  <th>Prod</th>\n'
#     html += '  <th>Hour</th>\n'
#     html += '  <th>Move</th>\n'
#     html += '  <th>Prod</th>\n'
#     html += '</tr>\n</thead>\n'
    
#     # Body rows
#     html += '<tbody>\n'
    
#     for day_report in report_data['daily_reports']:
#         date = day_report['date']
#         equipment_list = day_report['equipment']
        
#         for eq_idx, equipment in enumerate(equipment_list):
#             eq_name = equipment['equipment']
            
#             # Determine row styling
#             if eq_idx == 0:
#                 row_style = 'style="border-top: 2px solid #666;"'
#             else:
#                 row_style = ''
            
#             html += f'<tr {row_style}>\n'
            
#             # Date cell (only for first equipment)
#             if eq_idx == 0:
#                 html += f'  <td rowspan="{len(equipment_list)}" style="font-weight: bold; background-color: #f0f0f0;">{date}</td>\n'
            
#             # Equipment cell
#             html += f'  <td style="font-weight: bold; background-color: #f9f9f9;">{eq_name}</td>\n'
            
#             # Morning shift
#             morning = equipment.get('morning') or {}  # ✅ FIX: Handle None
#             if morning is not None:  # ✅ Check for None, not truthiness
#                 morning_hour = morning.get('crane_on_minute', 0) / 60
#                 morning_move = morning.get('number_of_move', '-')
#                 morning_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else '-'
#                 html += f'  <td style="background-color: #e3f2fd; text-align: center;">{morning_hour:.2f}</td>\n'
#                 html += f'  <td style="background-color: #e3f2fd; text-align: center;">{morning_move}</td>\n'
#                 html += f'  <td style="background-color: #e3f2fd; text-align: center;">{morning_prod}</td>\n'
#             else:
#                 # ✅ EXPLICITLY OUTPUT BLANK CELLS
#                 html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
            
#             # Night shift
#             night = equipment.get('night') or {}  # ✅ FIX: Handle None
#             if night is not None:  # ✅ Check for None, not truthiness
#                 night_hour = night.get('crane_on_minute', 0) / 60
#                 night_move = night.get('number_of_move', '-')
#                 night_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else '-'
#                 html += f'  <td style="background-color: #f5f5f5; text-align: center;">{night_hour:.2f}</td>\n'
#                 html += f'  <td style="background-color: #f5f5f5; text-align: center;">{night_move}</td>\n'
#                 html += f'  <td style="background-color: #f5f5f5; text-align: center;">{night_prod}</td>\n'
#             else:
#                 # ✅ EXPLICITLY OUTPUT BLANK CELLS
#                 html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
            
#             # Total
#             total = equipment.get('total') or {}  # ✅ FIX: Handle None
#             if total is not None:  # ✅ Check for None, not truthiness
#                 total_hour = total.get('crane_on_minute', 0) / 60
#                 total_move = total.get('number_of_move', '-')
#                 total_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else '-'
#                 html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{total_hour:.2f}</td>\n'
#                 html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{total_move}</td>\n'
#                 html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{total_prod}</td>\n'
#             else:
#                 html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
#                 html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
            
#             html += '</tr>\n'
    
#     html += '</tbody>\n</table>\n'
    
#     return html

def json_to_html_table(report_data):
    """
    Convert productivity report JSON to HTML table format
    WITH SUMMARY ROW
    
    Args:
        report_data: Dict from get_productivity_report_daily response
    
    Returns:
        HTML string of formatted table
    """
    
    if not report_data.get('daily_reports') or len(report_data['daily_reports']) == 0:
        return '<p>No data available</p>'
    
    html = '<table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">\n'
    
    # Header row
    html += '<thead>\n<tr style="background-color: #2c3e50; color: white;">\n'
    html += '  <th>Date</th>\n'
    html += '  <th>Equipment</th>\n'
    html += '  <th colspan="3">Morning (08:00-20:00)</th>\n'
    html += '  <th colspan="3">Night (20:00-08:00)</th>\n'
    html += '  <th colspan="3">Total</th>\n'
    html += '</tr>\n<tr style="background-color: #2c3e50; color: white;">\n'
    html += '  <th></th>\n'
    html += '  <th></th>\n'
    html += '  <th>Hour</th>\n'
    html += '  <th>Move</th>\n'
    html += '  <th>Prod</th>\n'
    html += '  <th>Hour</th>\n'
    html += '  <th>Move</th>\n'
    html += '  <th>Prod</th>\n'
    html += '  <th>Hour</th>\n'
    html += '  <th>Move</th>\n'
    html += '  <th>Prod</th>\n'
    html += '</tr>\n</thead>\n'
    
    # Body rows
    html += '<tbody>\n'
    
    # ✅ Track summary totals
    summary_morning_hour = 0
    summary_morning_move = 0
    summary_morning_prod_values = []
    summary_night_hour = 0
    summary_night_move = 0
    summary_night_prod_values = []
    summary_total_hour = 0
    summary_total_move = 0
    summary_total_prod_values = []
    
    for day_report in report_data['daily_reports']:
        date = day_report['date']
        equipment_list = day_report['equipment']
        
        for eq_idx, equipment in enumerate(equipment_list):
            eq_name = equipment['equipment']
            
            # Determine row styling
            if eq_idx == 0:
                row_style = 'style="border-top: 2px solid #666;"'
            else:
                row_style = ''
            
            html += f'<tr {row_style}>\n'
            
            # Date cell (only for first equipment)
            if eq_idx == 0:
                html += f'  <td rowspan="{len(equipment_list)}" style="font-weight: bold; background-color: #f0f0f0;">{date}</td>\n'
            
            # Equipment cell
            html += f'  <td style="font-weight: bold; background-color: #f9f9f9;">{eq_name}</td>\n'
            
            # Morning shift
            morning = equipment.get('morning') or {}
            if morning:
                morning_hour = morning.get('crane_on_minute', 0) / 60
                morning_move =int( morning.get('number_of_move', 0))
                morning_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else '-'
                
                # ✅ Accumulate for summary
                summary_morning_hour += morning_hour
                summary_morning_move += morning_move
                if morning_prod != '-':
                    summary_morning_prod_values.append(morning_prod)
                
                html += f'  <td style="background-color: #e3f2fd; text-align: center;">{morning_hour:.2f}</td>\n'
                html += f'  <td style="background-color: #e3f2fd; text-align: center;">{int(morning_move)}</td>\n'
                html += f'  <td style="background-color: #e3f2fd; text-align: center;">{morning_prod}</td>\n'
            else:
                html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #e3f2fd; text-align: center;">-</td>\n'
            
            # Night shift
            night = equipment.get('night') or {}
            if night:
                night_hour = night.get('crane_on_minute', 0) / 60
                night_move = int(night.get('number_of_move', 0))
                night_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else '-'
                
                # ✅ Accumulate for summary
                summary_night_hour += night_hour
                summary_night_move += night_move
                if night_prod != '-':
                    summary_night_prod_values.append(night_prod)
                
                html += f'  <td style="background-color: #f5f5f5; text-align: center;">{night_hour:.2f}</td>\n'
                html += f'  <td style="background-color: #f5f5f5; text-align: center;">{int(night_move)}</td>\n'
                html += f'  <td style="background-color: #f5f5f5; text-align: center;">{night_prod}</td>\n'
            else:
                html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #f5f5f5; text-align: center;">-</td>\n'
            
            # Total
            total = equipment.get('total') or {}
            if total:
                total_hour = total.get('crane_on_minute', 0) / 60
                total_move = int(total.get('number_of_move', 0))
                total_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else '-'
                
                # ✅ Accumulate for summary
                summary_total_hour += total_hour
                summary_total_move += total_move
                if total_prod != '-':
                    summary_total_prod_values.append(total_prod)
                
                html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{total_hour:.2f}</td>\n'
                html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{int(total_move)}</td>\n'
                html += f'  <td style="background-color: #fff9e6; text-align: center; font-weight: bold;">{total_prod}</td>\n'
            else:
                html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
                html += '  <td style="background-color: #fff9e6; text-align: center;">-</td>\n'
            
            html += '</tr>\n'
    
    # ✅ ADD SUMMARY ROW
    html += '<tr style="background-color: #4CAF50; color: white; font-weight: bold;">\n'
    html += '  <td colspan="2" style="text-align: center; padding: 12px;">📊 SUMMARY TOTAL</td>\n'
    
    # Morning summary
    # summary_morning_prod = round(sum(summary_morning_prod_values) / len(summary_morning_prod_values), 2) if summary_morning_prod_values else '-'
    summary_morning_prod = round(int(summary_morning_move)/summary_morning_hour  , 2) if summary_morning_prod_values else '-'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_morning_hour:.2f}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{int(summary_morning_move)}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_morning_prod}</td>\n'
    
    # Night summary
    # summary_night_prod = round(sum(summary_night_prod_values) / len(summary_night_prod_values), 2) if summary_night_prod_values else '-'
    summary_night_prod = round(int(summary_night_move)/summary_night_hour , 2) if summary_night_prod_values else '-'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_night_hour:.2f}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{int(summary_night_move)}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_night_prod}</td>\n'
    
    # Total summary
    summary_total_prod = round(int(summary_total_move)/summary_total_hour , 2) if summary_total_prod_values else '-'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_total_hour:.2f}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{int(summary_total_move)}</td>\n'
    html += f'  <td style="text-align: center; padding: 12px;">{summary_total_prod}</td>\n'
    
    html += '</tr>\n'
    
    html += '</tbody>\n</table>\n'
    
    return html


def json_to_plain_text_table(report_data):
    """
    Convert productivity report JSON to plain text table format (for console/email)
    
    Args:
        report_data: Dict from get_productivity_report_daily response
    
    Returns:
        Plain text formatted table string
    """
    
    if not report_data.get('daily_reports') or len(report_data['daily_reports']) == 0:
        return 'No data available'
    
    lines = []
    
    # Header
    lines.append('┌──────────┬───────────┬──────────────────┬──────────────────┬──────────────────┐')
    lines.append('│   Date   │ Equipment │ Morning          │ Night            │ Total            │')
    lines.append('│          │           │ Hr | Mv | Prod   │ Hr | Mv | Prod   │ Hr | Mv | Prod   │')
    lines.append('├──────────┼───────────┼──────────────────┼──────────────────┼──────────────────┤')
    
    first_date = True
    
    for day_report in report_data['daily_reports']:
        date = day_report['date']
        equipment_list = day_report['equipment']
        
        for eq_idx, equipment in enumerate(equipment_list):
            eq_name = equipment['equipment']
            
            # Date column (only for first equipment)
            if eq_idx == 0:
                date_str = date
                if not first_date:
                    lines.append('├──────────┼───────────┼──────────────────┼──────────────────┼──────────────────┤')
            else:
                date_str = '          '
            
            first_date = False
            
            # Morning ✅ FIX: Handle None
            morning = equipment.get('morning') or {}
            if morning:
                m_hr = morning.get('crane_on_minute', 0) / 60
                m_mv = morning.get('number_of_move', 0)
                m_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else 0
            else:
                m_hr = m_mv = m_prod = '-'
            
            # Night ✅ FIX: Handle None
            night = equipment.get('night') or {}
            if night:
                n_hr = night.get('crane_on_minute', 0) / 60
                n_mv = night.get('number_of_move', 0)
                n_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else 0
            else:
                n_hr = n_mv = n_prod = '-'
            
            # Total ✅ FIX: Handle None
            total = equipment.get('total') or {}
            if total:
                t_hr = total.get('crane_on_minute', 0) / 60
                t_mv = total.get('number_of_move', 0)
                t_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else 0
            else:
                t_hr = t_mv = t_prod = '-'
            
            # Format row
            morning_str = f'{m_hr} | {m_mv} | {m_prod}'
            night_str = f'{n_hr} | {n_mv} | {n_prod}'
            total_str = f'{t_hr} | {t_mv} | {t_prod}'
            
            line = f'│{date_str}│{eq_name:^11}│{morning_str:^18}│{night_str:^18}│{total_str:^18}│'
            lines.append(line)
    
    lines.append('└──────────┴───────────┴──────────────────┴──────────────────┴──────────────────┘')
    
    return '\n'.join(lines)


def json_to_csv(report_data):
    """
    Convert productivity report JSON to CSV format
    
    Args:
        report_data: Dict from get_productivity_report_daily response
    
    Returns:
        CSV formatted string
    """
    
    if not report_data.get('daily_reports') or len(report_data['daily_reports']) == 0:
        return ''
    
    lines = []
    lines.append('Date,Equipment,Morning Hour,Morning Move,Morning Productivity,Night Hour,Night Move,Night Productivity,Total Hour,Total Move,Total Productivity')
    
    for day_report in report_data['daily_reports']:
        date = day_report['date']
        equipment_list = day_report['equipment']
        
        for equipment in equipment_list:
            eq_name = equipment['equipment']
            
            # Morning ✅ FIX: Handle None
            morning = equipment.get('morning') or {}
            if morning:
                m_hr = morning.get('crane_on_minute', 0) / 60
                m_mv = morning.get('number_of_move', 0)
                m_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else ''
            else:
                m_hr = m_mv = m_prod = ''
            
            # Night ✅ FIX: Handle None
            night = equipment.get('night') or {}
            if night:
                n_hr = night.get('crane_on_minute', 0) / 60
                n_mv = night.get('number_of_move', 0)
                n_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else ''
            else:
                n_hr = n_mv = n_prod = ''
            
            # Total ✅ FIX: Handle None
            total = equipment.get('total') or {}
            if total:
                t_hr = total.get('crane_on_minute', 0) / 60
                t_mv = total.get('number_of_move', 0)
                t_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else ''
            else:
                t_hr = t_mv = t_prod = ''
            
            line = f'{date},"{eq_name}",{m_hr},{m_mv},{m_prod},{n_hr},{n_mv},{n_prod},{t_hr},{t_mv},{t_prod}'
            lines.append(line)
    
    return '\n'.join(lines)

from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# def json_to_excel(report_data):
#     """
#     Convert productivity report JSON to Excel format
    
#     Args:
#         report_data: Dict from get_productivity_report_daily response
    
#     Returns:
#         Bytes of Excel file
#     """
    
#     # Create workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Productivity Report"
    
#     # Define styles
#     header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
#     header_font = Font(bold=True, color="FFFFFF", size=12)
#     header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
#     title_font = Font(bold=True, size=14, color="2C3E50")
#     info_font = Font(size=11)
    
#     date_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
#     date_font = Font(bold=True)
    
#     equipment_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
#     equipment_font = Font(bold=True)
    
#     morning_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
#     night_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
#     total_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
#     total_font = Font(bold=True)
    
#     center_alignment = Alignment(horizontal="center", vertical="center")
#     left_alignment = Alignment(horizontal="left", vertical="center")
    
#     thin_border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Title
#     ws['A1'] = "Daily Productivity Report"
#     ws['A1'].font = title_font
#     ws.merge_cells('A1:K1')
    
#     # Info section
#     current_row = 3
#     ws[f'A{current_row}'] = "Date Range:"
#     ws[f'B{current_row}'] = f"{report_data['date_range']['from']} to {report_data['date_range']['to']}"
#     ws[f'A{current_row}'].font = Font(bold=True)
    
#     current_row += 1
#     ws[f'A{current_row}'] = "Equipment:"
#     ws[f'B{current_row}'] = report_data['summary']['total_equipment']
#     ws[f'A{current_row}'].font = Font(bold=True)
    
#     current_row += 1
#     ws[f'A{current_row}'] = "Generated:"
#     ws[f'B{current_row}'] = report_data.get('timestamp', 'N/A')
#     ws[f'A{current_row}'].font = Font(bold=True)
    
#     # Header row
#     current_row = 7
#     headers = [
#         'Date', 'Equipment',
#         'Morning Hour', 'Morning Move', 'Morning Productivity',
#         'Night Hour', 'Night Move', 'Night Productivity',
#         'Total Hour', 'Total Move', 'Total Productivity'
#     ]
    
#     for col_idx, header in enumerate(headers, start=1):
#         cell = ws.cell(row=current_row, column=col_idx)
#         cell.value = header
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = header_alignment
#         cell.border = thin_border
    
#     # Data rows
#     current_row = 8
    
#     for day_report in report_data['daily_reports']:
#         date = day_report['date']
#         equipment_list = day_report['equipment']
        
#         for eq_idx, equipment in enumerate(equipment_list):
#             eq_name = equipment['equipment']
            
#             # Date cell
#             date_cell = ws.cell(row=current_row, column=1)
#             date_cell.value = date if eq_idx == 0 else ""
#             date_cell.fill = date_fill
#             date_cell.font = date_font
#             date_cell.alignment = center_alignment
#             date_cell.border = thin_border
            
#             # Equipment cell
#             eq_cell = ws.cell(row=current_row, column=2)
#             eq_cell.value = eq_name
#             eq_cell.fill = equipment_fill
#             eq_cell.font = equipment_font
#             eq_cell.alignment = left_alignment
#             eq_cell.border = thin_border
            
#             col = 3

#             # Morning shift
#             morning = equipment.get('morning') or {}
#             if morning:
#                 m_hr = morning.get('crane_on_minute', 0) / 60
#                 m_mv = morning.get('number_of_move', 0)
#                 m_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else None
#             else:
#                 m_hr = m_mv = m_prod = None
            
#             for val in [m_hr, m_mv, m_prod]:
#                 cell = ws.cell(row=current_row, column=col)
#                 cell.value = val if val is not None else ""
#                 cell.fill = morning_fill
#                 cell.alignment = center_alignment
#                 cell.border = thin_border
#                 if val is not None and isinstance(val, (int, float)):
#                     cell.number_format = '0.00' if isinstance(val, float) else '0'
#                 col += 1
            
#             # Night shift
#             night = equipment.get('night') or {}
#             if night:
#                 n_hr = night.get('crane_on_minute', 0) / 60
#                 n_mv = night.get('number_of_move', 0)
#                 n_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else None
#             else:
#                 n_hr = n_mv = n_prod = None
            
#             for val in [n_hr, n_mv, n_prod]:
#                 cell = ws.cell(row=current_row, column=col)
#                 cell.value = val if val is not None else ""
#                 cell.fill = night_fill
#                 cell.alignment = center_alignment
#                 cell.border = thin_border
#                 if val is not None and isinstance(val, (int, float)):
#                     cell.number_format = '0.00' if isinstance(val, float) else '0'
#                 col += 1
            
#             # Total
#             total = equipment.get('total') or {}
#             if total:
#                 t_hr = total.get('crane_on_minute', 0) / 60
#                 t_mv = total.get('number_of_move', 0)
#                 t_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else None
#             else:
#                 t_hr = t_mv = t_prod = None
            
#             for val in [t_hr, t_mv, t_prod]:
#                 cell = ws.cell(row=current_row, column=col)
#                 cell.value = val if val is not None else ""
#                 cell.fill = total_fill
#                 cell.font = total_font
#                 cell.alignment = center_alignment
#                 cell.border = thin_border
#                 if val is not None and isinstance(val, (int, float)):
#                     cell.number_format = '0.00' if isinstance(val, float) else '0'
#                 col += 1
            
#             current_row += 1
    
#     # Adjust column widths
#     ws.column_dimensions['A'].width = 12
#     ws.column_dimensions['B'].width = 12
#     for col in range(3, 12):
#         ws.column_dimensions[get_column_letter(col)].width = 15
    
#     # Set row height for header
#     ws.row_dimensions[7].height = 25
    
#     # Save to bytes
#     excel_buffer = io.BytesIO()
#     wb.save(excel_buffer)
#     excel_buffer.seek(0)
    
#     return excel_buffer.getvalue()

def json_to_excel(report_data):
    """
    Convert productivity report JSON to Excel format
    WITH SUMMARY ROW
    
    Args:
        report_data: Dict from get_productivity_report_daily response
    
    Returns:
        Bytes of Excel file
    """
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productivity Report"
    
    # Define styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    title_font = Font(bold=True, size=14, color="2C3E50")
    info_font = Font(size=11)
    
    date_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    date_font = Font(bold=True)
    
    equipment_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    equipment_font = Font(bold=True)
    
    morning_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    night_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    total_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    total_font = Font(bold=True)
    
    # ✅ SUMMARY ROW STYLE
    summary_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    summary_font = Font(bold=True, color="FFFFFF", size=12)
    summary_alignment = Alignment(horizontal="center", vertical="center")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Daily Productivity Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:K1')
    
    # Info section
    current_row = 3
    ws[f'A{current_row}'] = "Date Range:"
    ws[f'B{current_row}'] = f"{report_data['date_range']['from']} to {report_data['date_range']['to']}"
    ws[f'A{current_row}'].font = Font(bold=True)
    
    current_row += 1
    ws[f'A{current_row}'] = "Equipment:"
    ws[f'B{current_row}'] = report_data['summary']['total_equipment']
    ws[f'A{current_row}'].font = Font(bold=True)
    
    current_row += 1
    ws[f'A{current_row}'] = "Generated:"
    ws[f'B{current_row}'] = report_data.get('timestamp', 'N/A')
    ws[f'A{current_row}'].font = Font(bold=True)
    
    # Header row
    current_row = 7
    headers = [
        'Date', 'Equipment',
        'Morning Hour', 'Morning Move', 'Morning Productivity',
        'Night Hour', 'Night Move', 'Night Productivity',
        'Total Hour', 'Total Move', 'Total Productivity'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    current_row = 8
    
    # ✅ Track summary totals
    summary_morning_hour = 0
    summary_morning_move = 0
    summary_morning_prod_values = []
    summary_night_hour = 0
    summary_night_move = 0
    summary_night_prod_values = []
    summary_total_hour = 0
    summary_total_move = 0
    summary_total_prod_values = []
    
    for day_report in report_data['daily_reports']:
        date = day_report['date']
        equipment_list = day_report['equipment']
        
        for eq_idx, equipment in enumerate(equipment_list):
            eq_name = equipment['equipment']
            
            # Date cell
            date_cell = ws.cell(row=current_row, column=1)
            date_cell.value = date if eq_idx == 0 else ""
            date_cell.fill = date_fill
            date_cell.font = date_font
            date_cell.alignment = center_alignment
            date_cell.border = thin_border
            
            # Equipment cell
            eq_cell = ws.cell(row=current_row, column=2)
            eq_cell.value = eq_name
            eq_cell.fill = equipment_fill
            eq_cell.font = equipment_font
            eq_cell.alignment = left_alignment
            eq_cell.border = thin_border
            
            col = 3

            # Morning shift
            morning = equipment.get('morning') or {}
            if morning:
                m_hr = morning.get('crane_on_minute', 0) / 60
                m_mv = morning.get('number_of_move', 0)
                m_prod = round(morning.get('productivity', 0), 2) if morning.get('productivity') else None
                
                # ✅ Accumulate
                summary_morning_hour += m_hr
                summary_morning_move += m_mv
                if m_prod is not None:
                    summary_morning_prod_values.append(m_prod)
            else:
                m_hr = m_mv = m_prod = None
            
            for val in [m_hr, int(m_mv) if m_mv is not None else None, m_prod]:
                cell = ws.cell(row=current_row, column=col)
                cell.value = val if val is not None else ""
                cell.fill = morning_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                if val is not None and isinstance(val, (int, float)):
                    cell.number_format = '0.00' if isinstance(val, float) else '0'
                col += 1
            
            # Night shift
            night = equipment.get('night') or {}
            if night:
                n_hr = night.get('crane_on_minute', 0) / 60
                n_mv = night.get('number_of_move', 0)
                n_prod = round(night.get('productivity', 0), 2) if night.get('productivity') else None
                
                # ✅ Accumulate
                summary_night_hour += n_hr
                summary_night_move += n_mv
                if n_prod is not None:
                    summary_night_prod_values.append(n_prod)
            else:
                n_hr = n_mv = n_prod = None
            
            for val in [n_hr, int(n_mv) if n_mv is not None else None, n_prod]:
                cell = ws.cell(row=current_row, column=col)
                cell.value = val if val is not None else ""
                cell.fill = night_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                if val is not None and isinstance(val, (int, float)):
                    cell.number_format = '0.00' if isinstance(val, float) else '0'
                col += 1
            
            # Total
            total = equipment.get('total') or {}
            if total:
                t_hr = total.get('crane_on_minute', 0) / 60
                t_mv = total.get('number_of_move', 0)
                t_prod = round(total.get('productivity', 0), 2) if total.get('productivity') else None
                
                # ✅ Accumulate
                summary_total_hour += t_hr
                summary_total_move += t_mv
                if t_prod is not None:
                    summary_total_prod_values.append(t_prod)
            else:
                t_hr = t_mv = t_prod = None
            
            for val in [t_hr, int(t_mv) if t_mv is not None else None, t_prod]:
                cell = ws.cell(row=current_row, column=col)
                cell.value = val if val is not None else ""
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = center_alignment
                cell.border = thin_border
                if val is not None and isinstance(val, (int, float)):
                    cell.number_format = '0.00' if isinstance(val, float) else '0'
                col += 1
            
            current_row += 1
    
    # ✅ ADD SUMMARY ROW
    summary_row = current_row
    
    # Merge cells for label
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    label_cell = ws.cell(row=summary_row, column=1)
    label_cell.value = "📊 SUMMARY TOTAL"
    label_cell.fill = summary_fill
    label_cell.font = summary_font
    label_cell.alignment = summary_alignment
    label_cell.border = thin_border
    
    col = 3
    
    # Morning summary
    # summary_morning_prod = round(sum(summary_morning_prod_values) / len(summary_morning_prod_values), 2) if summary_morning_prod_values else None
    summary_morning_prod = round(int(summary_morning_move)/summary_morning_hour, 2) if summary_morning_prod_values else None
    for val in [summary_morning_hour, int(summary_morning_move), summary_morning_prod]:
        cell = ws.cell(row=summary_row, column=col)
        cell.value = val
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = summary_alignment
        cell.border = thin_border
        if isinstance(val, float):
            cell.number_format = '0.00'
        col += 1
    
    # Night summary
    summary_night_prod = round(int(summary_night_move)/summary_night_hour, 2) if summary_night_prod_values else None
    for val in [summary_night_hour, int(summary_night_move), summary_night_prod]:
        cell = ws.cell(row=summary_row, column=col)
        cell.value = val
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = summary_alignment
        cell.border = thin_border
        if isinstance(val, float):
            cell.number_format = '0.00'
        col += 1
    
    # Total summary
    summary_total_prod = round(int(summary_total_move)/summary_total_hour, 2) if summary_total_prod_values else None
    for val in [summary_total_hour, int(summary_total_move), summary_total_prod]:
        cell = ws.cell(row=summary_row, column=col)
        cell.value = val
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = summary_alignment
        cell.border = thin_border
        if isinstance(val, float):
            cell.number_format = '0.00'
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    for col in range(3, 12):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Set row height for header
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[summary_row].height = 25
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()