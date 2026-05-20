"""
Item Summary Report HTML and Excel Generators
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging

logger = logging.getLogger(__name__)


def item_summary_to_html(report_data):
    """
    Convert item summary report to HTML table
    
    Rows: Items
    Columns: Equipment
    Values: Item values
    """
    
    items = report_data.get('items', [])
    equipment = report_data.get('equipment', [])
    data = report_data.get('data', {})
    period_label = report_data.get('period_label', 'N/A')
    date_range = report_data.get('date_range', {})
    
    if not items or not equipment:
        return '<p>No data available</p>'
    
    html = '<table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%;">\n'
    
    # Title row
    html += f'<thead>\n'
    html += f'<tr style="background-color: #2c3e50; color: white;">\n'
    html += f'  <th colspan="{len(equipment) + 1}" style="text-align: center; padding: 12px; font-size: 14px;">📊 Item Summary Report</th>\n'
    html += f'</tr>\n'
    
    # Info row
    html += f'<tr style="background-color: #ecf0f1;">\n'
    html += f'  <td colspan="{len(equipment) + 1}" style="padding: 8px;">\n'
    html += f'    <strong>Period:</strong> {period_label} | '
    html += f'    <strong>Date Range:</strong> {date_range.get("from", "N/A")} to {date_range.get("to", "N/A")}\n'
    html += f'  </td>\n'
    html += f'</tr>\n'
    
    # Header row with equipment names
    html += f'<tr style="background-color: #34495e; color: white;">\n'
    html += f'  <th style="text-align: center; padding: 10px;">Item Name</th>\n'
    
    for eq in equipment:
        html += f'  <th style="text-align: center; padding: 10px;">{eq}</th>\n'
    
    html += f'</tr>\n'
    html += f'</thead>\n'
    
    # Data rows
    html += f'<tbody>\n'
    
    for idx, item_name in enumerate(items):
        # Alternate row colors
        bg_color = '#f8f9fa' if idx % 2 == 0 else '#ffffff'
        
        html += f'<tr style="background-color: {bg_color};">\n'
        html += f'  <td style="font-weight: bold; padding: 10px;">{item_name}</td>\n'
        
        # Values for each equipment
        for eq in equipment:
            value = data.get(item_name, {}).get(eq, 0)
            
            # Format value
            if isinstance(value, float):
                formatted_value = f'{value:.2f}'
            else:
                formatted_value = str(value)
            
            html += f'  <td style="text-align: center; padding: 10px;">{formatted_value}</td>\n'
        
        html += f'</tr>\n'
    
    html += f'</tbody>\n'
    html += f'</table>\n'
    
    return html


def item_summary_to_excel(report_data):
    """
    Convert item summary report to Excel file
    
    Rows: Items
    Columns: Equipment
    Values: Item values
    
    Returns: Excel file bytes
    """
    
    items = report_data.get('items', [])
    equipment = report_data.get('equipment', [])
    data = report_data.get('data', {})
    period_label = report_data.get('period_label', 'N/A')
    date_range = report_data.get('date_range', {})
    timestamp = report_data.get('timestamp', 'N/A')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Summary"
    
    # Define styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_font = Font(size=10)
    info_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    
    item_font = Font(bold=True)
    item_alignment = Alignment(horizontal="left", vertical="center")
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ✅ ROW 1: Title
    current_row = 1
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(equipment) + 1)}{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "📊 Item Summary Report"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    ws.row_dimensions[current_row].height = 25
    
    # ✅ ROW 2: Info
    current_row = 2
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(equipment) + 1)}{current_row}')
    info_cell = ws.cell(row=current_row, column=1)
    info_cell.value = (
        f"Period: {period_label} | "
        f"Date Range: {date_range.get('from', 'N/A')} to {date_range.get('to', 'N/A')} | "
        f"Generated: {timestamp}"
    )
    info_cell.font = info_font
    info_cell.fill = info_fill
    info_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[current_row].height = 30
    
    # ✅ ROW 4: Headers (skip row 3)
    current_row = 4
    
    # Item name header
    item_header = ws.cell(row=current_row, column=1)
    item_header.value = "Item Name"
    item_header.font = header_font
    item_header.fill = header_fill
    item_header.alignment = header_alignment
    item_header.border = thin_border
    
    # Equipment headers
    for col_idx, eq in enumerate(equipment, start=2):
        eq_header = ws.cell(row=current_row, column=col_idx)
        eq_header.value = eq
        eq_header.font = header_font
        eq_header.fill = header_fill
        eq_header.alignment = header_alignment
        eq_header.border = thin_border
    
    ws.row_dimensions[current_row].height = 20
    
    # ✅ DATA ROWS
    current_row = 5
    
    for idx, item_name in enumerate(items):
        # Item name cell
        item_cell = ws.cell(row=current_row, column=1)
        item_cell.value = item_name
        item_cell.font = item_font
        item_cell.alignment = item_alignment
        item_cell.border = thin_border
        
        # Data cells
        for col_idx, eq in enumerate(equipment, start=2):
            value = data.get(item_name, {}).get(eq, 0)
            
            data_cell = ws.cell(row=current_row, column=col_idx)
            data_cell.value = value if value != 0 else ""
            data_cell.alignment = data_alignment
            data_cell.border = thin_border
            
            # Format number
            if isinstance(value, float):
                data_cell.number_format = '0.00'
            elif isinstance(value, int):
                data_cell.number_format = '0'
        
        current_row += 1
    
    # ✅ COLUMN WIDTHS
    ws.column_dimensions['A'].width = 25
    for col in range(2, len(equipment) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # ✅ SAVE TO BYTES
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    logger.info(f"✅ Excel generated with {len(items)} items and {len(equipment)} equipment")
    
    return excel_buffer.getvalue()